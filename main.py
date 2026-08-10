
import os
import math
import shutil
import re
import itertools
import time
from pathlib import Path
from collections import Counter
from itertools import combinations

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from mpl_toolkits.axes_grid1 import make_axes_locatable
from scipy.spatial.distance import pdist, squareform
from scipy.cluster.hierarchy import linkage, dendrogram
from sklearn.manifold import TSNE
from sklearn.metrics import pairwise_distances
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


INPUT_XLSX = 'data.xlsx'
OUTPUT_DIR = Path('output')
OUTPUT_DIR.mkdir(exist_ok=True)

# Global parameters
TOP_N_PAIRS = 50
TOP_N_PAIR_HEATMAP = 20
TOP_N_PESTICIDES = 10
TOP_N_RULES = 50
MIN_SUPPORT = 0.1
MIN_CONFIDENCE = 0.5



def ensure_dir(path):
    """Create a directory path if it does not already exist.

    Parameters
    ----------
    path : str or Path
        Directory path to create.
    """
    Path(path).mkdir(parents=True, exist_ok=True)


def sanitize_sheet(name, suffix=''):
    """Trim and adapt a sheet name to Excel's 31-character limit.

    Parameters
    ----------
    name : str
        Base sheet name.
    suffix : str, optional
        Optional suffix appended while preserving the 31-character limit.
    """
    base = str(name)
    if suffix:
        keep = 31 - len(suffix)
        return f"{base[:keep]}{suffix}"
    return base[:31]


def sanitize_filename(name):
    """Convert text to a filesystem-safe file-name fragment.

    Parameters
    ----------
    name : str
        Raw text to sanitize for filenames.
    """
    return ''.join(ch if ch.isalnum() or ch in ['_', '-'] else '_' for ch in str(name))


def esc_tex(s):
    """Escape special characters for LaTeX output.

    Parameters
    ----------
    s : Any
        Value converted to string and escaped for LaTeX.
    """
    s = str(s)
    repl = {
        '\\': r'\textbackslash{}',
        '&': r'\&', '%': r'\%', '$': r'\$', '#': r'\#',
        '_': r'\_', '{': r'\{', '}': r'\}', '~': r'\textasciitilde{}',
        '^': r'\textasciicircum{}',
    }
    for k, v in repl.items():
        s = s.replace(k, v)
    return s


def fmt_num(x, decimals=3):
    """Format numeric values for LaTeX tables, including infinity.

    Parameters
    ----------
    x : Any
        Value to format.
    decimals : int, optional
        Number of decimal places for finite numeric values.
    """
    if isinstance(x, float) and math.isinf(x):
        return r'$\infty$'
    try:
        return f'{float(x):.{decimals}f}'
    except (TypeError, ValueError):
        return esc_tex(x)


def load_workbook_sheets(xlsx_path=INPUT_XLSX):
    """Load all sheets from the pesticide workbook into cleaned DataFrames.

    Parameters
    ----------
    xlsx_path : str, optional
        Path to the source Excel workbook.
    """
    xl = pd.ExcelFile(xlsx_path)
    sheets = {}
    for name in xl.sheet_names:
        df = xl.parse(name)
        df = df.dropna(axis=1, how='all').dropna(axis=0, how='all').copy()
        df.rename(columns={df.columns[0]: 'Pesticid'}, inplace=True)
        df = df[df['Pesticid'].notna()].copy()
        sheets[name] = df
    return sheets


def to_binary_sheet(df):
    """Convert a pesticide sheet to binary presence/absence values.

    Parameters
    ----------
    df : pandas.DataFrame
        Input sheet with pesticide names in the first column and sampling columns after it.
    """
    out = df.copy()
    value_cols = [c for c in out.columns if c != 'Pesticid']
    for col in value_cols:
        out[col] = ((out[col].notna()) & (out[col].astype(str).str.strip() != '')).astype(int)
    return out


def build_binary_outputs(sheets):
    """Build unified binary matrices for all sheets and export them.

    Parameters
    ----------
    sheets : dict[str, pandas.DataFrame]
        Mapping from sheet name to raw sheet DataFrame.
    """
    all_pesticides = sorted({str(p) for df in sheets.values() for p in df['Pesticid'].astype(str).tolist()})
    binary_sheets = {}
    for name, df in sheets.items():
        b = to_binary_sheet(df).set_index('Pesticid')
        b = b.reindex(all_pesticides).fillna(0).astype(int)
        binary_sheets[name] = b

    with pd.ExcelWriter(OUTPUT_DIR / '01_binary_presence_by_sheet.xlsx', engine='openpyxl') as writer:
        for name, b in binary_sheets.items():
            b.to_excel(writer, sheet_name=sanitize_sheet(name))

    wide_parts = []
    for name, b in binary_sheets.items():
        t = b.copy()
        t.columns = pd.MultiIndex.from_product([[name], t.columns])
        wide_parts.append(t)
    wide_binary = pd.concat(wide_parts, axis=1)
    wide_binary.index.name = 'Pesticid'
    wide_binary.to_csv(OUTPUT_DIR / '02_binary_presence_all_sheets_wide.csv')
    return binary_sheets, wide_binary


def compute_tsne(binary_sheets):
    """Compute t-SNE coordinates from binary locality-month vectors.

    Parameters
    ----------
    binary_sheets : dict[str, pandas.DataFrame]
        Mapping from sheet name to binary pesticide-by-month matrix.
    """
    vectors, labels = [], []
    for sheet, df in binary_sheets.items():
        for col in df.columns:
            vectors.append(df[col].astype(int).values)
            labels.append((sheet, str(col)))
    X = np.vstack(vectors)
    D = pairwise_distances(X, metric='hamming')
    tsne = TSNE(n_components=2, metric='precomputed', random_state=42, perplexity=10, init='random')
    emb = tsne.fit_transform(D)
    out = pd.DataFrame({
        'sheet': [s for s, _ in labels],
        'month': [m for _, m in labels],
        'tsne1': emb[:, 0],
        'tsne2': emb[:, 1],
    })
    out.to_csv(OUTPUT_DIR / '03_tsne_hamming_coordinates.csv', index=False)

    plt.figure(figsize=(9, 7))
    for sheet, g in out.groupby('sheet'):
        plt.scatter(g['tsne1'], g['tsne2'], label=sheet, alpha=0.8, s=35)
    plt.xlabel('t-SNE 1')
    plt.ylabel('t-SNE 2')
    plt.legend(bbox_to_anchor=(1.02, 1), loc='upper left', fontsize=8)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / '04_tsne_hamming_scatterplot.pdf', bbox_inches='tight')
    plt.close()
    return out


def relative_frequency_outputs(sheets):
    """Compute pesticide relative frequencies by locality and related exports.

    Parameters
    ----------
    sheets : dict[str, pandas.DataFrame]
        Mapping from sheet name to raw sheet DataFrame.
    """
    rows = []
    for sheet, df in sheets.items():
        b = to_binary_sheet(df)
        value_cols = [c for c in b.columns if c != 'Pesticid']
        tmp = pd.DataFrame({
            'Locality': sheet,
            'Pesticide': b['Pesticid'].astype(str),
            'rel_freq': b[value_cols].sum(axis=1) / len(value_cols)
        })
        rows.append(tmp)
    result = pd.concat(rows, ignore_index=True)
    result.to_excel(OUTPUT_DIR / '05_relative_frequency_by_locality.xlsx', index=False)

    with pd.ExcelWriter(OUTPUT_DIR / '06_relative_frequency_by_locality_sheets.xlsx', engine='openpyxl') as writer:
        for loc in result['Locality'].unique():
            result[result['Locality'] == loc].to_excel(writer, sheet_name=sanitize_sheet(loc), index=False)

    pivot = result.pivot_table(index='Pesticide', columns='Locality', values='rel_freq')
    pivot.to_excel(OUTPUT_DIR / '07_relative_frequency_localities_wide.xlsx')
    pivot.corr().to_excel(OUTPUT_DIR / '08_relative_frequency_locality_correlation_matrix.xlsx')

    top10 = result.sort_values(['Locality', 'rel_freq', 'Pesticide'], ascending=[True, False, True]).groupby('Locality').head(TOP_N_PESTICIDES).copy()
    top10['Rank'] = top10.groupby('Locality').cumcount() + 1
    top10_matrix = top10.pivot(index='Rank', columns='Locality', values='Pesticide')
    top10_xlsx = OUTPUT_DIR / '09_top_pesticides_by_locality.xlsx'
    with pd.ExcelWriter(top10_xlsx, engine='openpyxl') as writer:
        top10_matrix.to_excel(writer, sheet_name='top10_pesticides')
        top10.to_excel(writer, sheet_name='long_format', index=False)

    basecolors = ['FFB3E6', 'FFB3BA', 'FFDFBA', 'FFFFBA', 'BAFFC9', 'BAE1FF', 'E2C6FF', 'FFD1DC', 'C4FAF8', 'F1F0C0']
    pesticides = pd.unique(top10_matrix.values.ravel('K'))
    pesticides = [p for p in pesticides if pd.notna(p)]
    colormap = {p: basecolors[i % len(basecolors)] for i, p in enumerate(pesticides)}
    for outname, distinct in [('10_top_pesticides_by_locality_colored.xlsx', False), ('11_top_pesticides_by_locality_colored_distinct.xlsx', True)]:
        shutil.copy(top10_xlsx, OUTPUT_DIR / outname)
        wb = load_workbook(OUTPUT_DIR / outname)
        ws = wb['top10_pesticides']
        if distinct:
            loc_cols = {c: {} for c in range(2, ws.max_column + 1)}
        for r in range(2, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val in colormap:
                    if distinct:
                        if val not in loc_cols[c]:
                            loc_cols[c][val] = basecolors[len(loc_cols[c]) % len(basecolors)]
                        color = loc_cols[c][val]
                    else:
                        color = colormap[val]
                    ws.cell(row=r, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        wb.save(OUTPUT_DIR / outname)

    g = top10.sort_values(['Locality', 'rel_freq'], ascending=[True, False]).copy()
    g['Locality'] = g['Locality'].str.replace('_', ' ', regex=False)
    n = g['Locality'].nunique()
    cols = 3
    rows_n = math.ceil(n / cols)
    fig, axes = plt.subplots(rows_n, cols, figsize=(16, 4.5 * rows_n), sharex=False)
    axes = np.array(axes).reshape(-1)
    palette = sns.color_palette('tab10', TOP_N_PESTICIDES)
    for ax, (loc, sub) in zip(axes, g.groupby('Locality')):
        ax.barh(sub['Pesticide'][::-1], sub['rel_freq'][::-1], color=palette[:len(sub)])
        ax.set_title(loc, fontsize=11)
        ax.set_xlabel('Relative frequency')
        ax.tick_params(axis='y', labelsize=8)
    for ax in axes[len(list(g.groupby('Locality'))):]:
        ax.axis('off')
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / '12_relative_frequency_top_pesticides_barplots.pdf', bbox_inches='tight')
    plt.close()
    return result, top10


def pair_and_triplet_itemsets(sheets):
    """Compute pair and triplet itemsets for each locality.

    Parameters
    ----------
    sheets : dict[str, pandas.DataFrame]
        Mapping from sheet name to raw sheet DataFrame.
    """
    all_results = []
    per_sheet = {}
    for sheet, df in sheets.items():
        b = to_binary_sheet(df)
        value_cols = [c for c in b.columns if c != 'Pesticid']
        pair_counts = Counter()
        triplet_counts = Counter()
        n_transactions = len(value_cols)
        for col in value_cols:
            present = sorted(b.loc[b[col] == 1, 'Pesticid'].astype(str).tolist())
            for a, b2 in itertools.combinations(present, 2):
                pair_counts[(a, b2)] += 1
            for a, b2, c in itertools.combinations(present, 3):
                triplet_counts[(a, b2, c)] += 1
        rows = []
        for (a, b2), cnt in pair_counts.items():
            rows.append({'Locality': sheet, 'ItemsetSize': 2, 'Pesticide1': a, 'Pesticide2': b2, 'SupportCount': cnt, 'SupportRel': cnt / n_transactions})
        for (a, b2, c), cnt in triplet_counts.items():
            rows.append({'Locality': sheet, 'ItemsetSize': 3, 'Pesticide1': a, 'Pesticide2': b2, 'Pesticide3': c, 'SupportCount': cnt, 'SupportRel': cnt / n_transactions})
        df_sheet = pd.DataFrame(rows)
        all_results.append(df_sheet)
        per_sheet[sheet] = df_sheet
    result = pd.concat(all_results, ignore_index=True)
    result.to_excel(OUTPUT_DIR / '13_itemsets_pairs_triplets_all_localities.xlsx', index=False)

    with pd.ExcelWriter(OUTPUT_DIR / '14_itemsets_pairs_triplets_split_by_locality.xlsx', engine='openpyxl') as writer:
        for loc, sub in per_sheet.items():
            sub.to_excel(writer, sheet_name=sanitize_sheet(loc), index=False)

    src = OUTPUT_DIR / '14_itemsets_pairs_triplets_split_by_locality.xlsx'
    dst = OUTPUT_DIR / '15_itemsets_pairs_and_triplets_separate_sheets.xlsx'
    shutil.copy(src, dst)
    wb = load_workbook(dst)
    for name in list(wb.sheetnames):
        ws = wb[name]
        data = list(ws.values)
        if not data:
            continue
        header, *rows = data
        if 'ItemsetSize' not in header:
            continue
        idx = header.index('ItemsetSize')
        pairs_rows = [header] + [row for row in rows if row[idx] == 2]
        trips_rows = [header] + [row for row in rows if row[idx] == 3]
        base = name[:23]
        for sfx, out_rows in [('_pairs', pairs_rows), ('_triplets', trips_rows)]:
            new_name = sanitize_sheet(base, sfx)
            if new_name in wb.sheetnames:
                del wb[new_name]
            nws = wb.create_sheet(new_name)
            for r_idx, row in enumerate(out_rows, start=1):
                for c_idx, val in enumerate(row, start=1):
                    nws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(dst)
    return result


def compute_pair_itemsets_confidence(sheets):
    """Compute pairwise association-rule metrics for each locality.

    Parameters
    ----------
    sheets : dict[str, pandas.DataFrame]
        Mapping from sheet name to raw sheet DataFrame.
    """
    all_rows = []
    per_locality_frames = {}
    for sheet_name, df in sheets.items():
        b = to_binary_sheet(df)
        pesticides = b['Pesticid'].astype(str).tolist()
        value_cols = [c for c in b.columns if c != 'Pesticid']
        presence = b[value_cols].astype(int).values
        transactions = []
        for m in range(len(value_cols)):
            t = [pesticides[i] for i in range(len(pesticides)) if presence[i, m] == 1]
            transactions.append(t)
        num_transactions = len(transactions)
        item_counts = Counter()
        pair_counts = Counter()
        for t in transactions:
            s = sorted(set(t))
            for it in s:
                item_counts[it] += 1
            for comb in combinations(s, 2):
                pair_counts[comb] += 1
        rows = []
        for (a, b2), count_ab in pair_counts.items():
            count_a = item_counts[a]
            count_b = item_counts[b2]
            support_ab = count_ab / num_transactions
            support_a = count_a / num_transactions
            support_b = count_b / num_transactions
            conf_a_to_b = count_ab / count_a if count_a else 0.0
            conf_b_to_a = count_ab / count_b if count_b else 0.0
            lift = (support_ab / (support_a * support_b)) if (support_a and support_b) else 0.0
            conv_a_to_b = float('inf') if conf_a_to_b == 1 else ((1 - support_b) / (1 - conf_a_to_b) if conf_a_to_b < 1 else float('inf'))
            conv_b_to_a = float('inf') if conf_b_to_a == 1 else ((1 - support_a) / (1 - conf_b_to_a) if conf_b_to_a < 1 else float('inf'))
            rows.append({
                'Locality': sheet_name,
                'Pesticide1': a,
                'Pesticide2': b2,
                'SupportCount_A': count_a,
                'SupportCount_B': count_b,
                'SupportCount_AB': count_ab,
                'NumTransactions': num_transactions,
                'SupportRel_AB': round(support_ab, 6),
                'Confidence_A_to_B': round(conf_a_to_b, 6),
                'Confidence_B_to_A': round(conf_b_to_a, 6),
                'Lift': round(lift, 6),
                'Conviction_A_to_B': conv_a_to_b if math.isinf(conv_a_to_b) else round(conv_a_to_b, 6),
                'Conviction_B_to_A': conv_b_to_a if math.isinf(conv_b_to_a) else round(conv_b_to_a, 6),
            })
        df_pairs = pd.DataFrame(rows).sort_values(['SupportRel_AB', 'Confidence_A_to_B'], ascending=[False, False]).reset_index(drop=True)
        per_locality_frames[sheet_name] = df_pairs
        all_rows.append(df_pairs)
    combined = pd.concat(all_rows, ignore_index=True)
    combined.to_csv(OUTPUT_DIR / '16_pair_rules_metrics_all_localities.csv', sep=';', index=False)
    with pd.ExcelWriter(OUTPUT_DIR / '17_pair_rules_metrics_by_locality.xlsx', engine='openpyxl') as writer:
        for name, dfp in per_locality_frames.items():
            dfp.drop(columns=['Locality']).to_excel(writer, sheet_name=sanitize_sheet(name), index=False)

    top50_path = OUTPUT_DIR / '18_pair_rules_topN_by_locality.xlsx'
    with pd.ExcelWriter(top50_path, engine='openpyxl') as writer:
        for name, dfp in per_locality_frames.items():
            dfp.head(TOP_N_PAIRS).drop(columns=['Locality']).to_excel(writer, sheet_name=sanitize_sheet(name), index=False)

    src = top50_path
    dst = OUTPUT_DIR / '19_pair_rules_topN_by_locality_report.xlsx'
    xls = pd.ExcelFile(src)
    with pd.ExcelWriter(dst, engine='openpyxl') as writer:
        for sheet in xls.sheet_names:
            df = pd.read_excel(src, sheet_name=sheet)
            df['SupportRel_A'] = df['SupportCount_A'] / df['NumTransactions']
            df['SupportRel_B'] = df['SupportCount_B'] / df['NumTransactions']
            df = df.drop(columns=['SupportCount_AB', 'SupportCount_A', 'SupportCount_B', 'NumTransactions'])
            df = df.rename(columns={'Pesticide1': 'Pesticide A', 'Pesticide2': 'Pesticide B'})
            ordered = ['Pesticide A', 'Pesticide B', 'SupportRel_A', 'SupportRel_B', 'SupportRel_AB', 'Confidence_A_to_B', 'Confidence_B_to_A', 'Lift', 'Conviction_A_to_B', 'Conviction_B_to_A']
            df = df[ordered]
            for c in ['SupportRel_A', 'SupportRel_B']:
                df[c] = df[c].round(6)
            df.insert(0, 'RowIndex', range(1, len(df) + 1))
            df.to_excel(writer, sheet_name=sanitize_sheet(sheet), index=False)
    return combined





def compute_association_rules_triplets(sheets, min_support=MIN_SUPPORT, min_confidence=MIN_CONFIDENCE):
    """Compute association rules for pesticide triplets with support/confidence filtering.

    Parameters
    ----------
    sheets : dict[str, pandas.DataFrame]
        Mapping from sheet name to raw sheet DataFrame.
    min_support : float, optional
        Minimum relative support threshold for retaining a rule.
    min_confidence : float, optional
        Minimum confidence threshold for retaining a rule.
    """
    all_rows = []
    per_locality_frames = {}

    for sheet_name, df in sheets.items():
        b = to_binary_sheet(df)
        pesticides = b['Pesticid'].astype(str).tolist()
        value_cols = [c for c in b.columns if c != 'Pesticid']
        presence = b[value_cols].astype(int).values

        transactions = []
        for m in range(len(value_cols)):
            t = sorted({pesticides[i] for i in range(len(pesticides)) if presence[i, m] == 1})
            transactions.append(t)

        num_transactions = len(transactions)
        item_counts = Counter()
        pair_counts = Counter()
        triplet_counts = Counter()

        for t in transactions:
            for it in t:
                item_counts[it] += 1
            for comb in combinations(t, 2):
                pair_counts[comb] += 1
            for comb in combinations(t, 3):
                triplet_counts[comb] += 1

        rows = []
        for (a, b2, c), count_abc in triplet_counts.items():
            support_abc = count_abc / num_transactions
            pair_ab = pair_counts.get(tuple(sorted((a, b2))), 0)
            pair_ac = pair_counts.get(tuple(sorted((a, c))), 0)
            pair_bc = pair_counts.get(tuple(sorted((b2, c))), 0)
            support_a = item_counts[a] / num_transactions
            support_b = item_counts[b2] / num_transactions
            support_c = item_counts[c] / num_transactions
            support_ab = pair_ab / num_transactions if num_transactions else 0.0
            support_ac = pair_ac / num_transactions if num_transactions else 0.0
            support_bc = pair_bc / num_transactions if num_transactions else 0.0

            triplet_rules = [
                ((a, b2), c, pair_ab, support_ab, support_c),
                ((a, c), b2, pair_ac, support_ac, support_b),
                ((b2, c), a, pair_bc, support_bc, support_a),
                ((a,), (b2, c), item_counts[a], support_a, support_bc),
                ((b2,), (a, c), item_counts[b2], support_b, support_ac),
                ((c,), (a, b2), item_counts[c], support_c, support_ab),
            ]

            for antecedent, consequent, antecedent_count, antecedent_support, consequent_support in triplet_rules:
                confidence = count_abc / antecedent_count if antecedent_count else 0.0
                if support_abc >= min_support and confidence >= min_confidence:
                    rows.append({
                        'Locality': sheet_name,
                        'RuleSize': 3,
                        'Antecedent': ' + '.join(antecedent),
                        'Consequent': consequent if isinstance(consequent, str) else ' + '.join(consequent),
                        'SupportRel': round(support_abc, 6),
                        'Confidence': round(confidence, 6),
                        'Lift': round((support_abc / (antecedent_support * consequent_support)) if (antecedent_support and consequent_support) else 0.0, 6),
                    })

        df_rules = pd.DataFrame(rows)
        if not df_rules.empty:
            df_rules = df_rules.sort_values(['Confidence', 'SupportRel', 'Antecedent', 'Consequent'], ascending=[False, False, True, True]).reset_index(drop=True)
        per_locality_frames[sheet_name] = df_rules
        all_rows.append(df_rules)

    combined = pd.concat(all_rows, ignore_index=True) if all_rows else pd.DataFrame()
    combined.to_csv(OUTPUT_DIR / '26_association_rules_triplets_all_localities.csv', sep=';', index=False)

    with pd.ExcelWriter(OUTPUT_DIR / '27_association_rules_triplets_by_locality.xlsx', engine='openpyxl') as writer:
        for name, dfr in per_locality_frames.items():
            dfr.to_excel(writer, sheet_name=sanitize_sheet(name), index=False)

    with pd.ExcelWriter(OUTPUT_DIR / '28_association_rules_triplets_topN_by_locality.xlsx', engine='openpyxl') as writer:
        for name, dfr in per_locality_frames.items():
            dfr.head(TOP_N_RULES).to_excel(writer, sheet_name=sanitize_sheet(name), index=False)

    return combined


def pair_heatmaps_and_latex(relfreq_df):
    """Create pair-support heatmaps and LaTeX tables for top pair rules.

    Parameters
    ----------
    relfreq_df : pandas.DataFrame
        Table with columns including Locality, Pesticide, and rel_freq.
    """
    rel_df = relfreq_df.copy()
    rel_df['LocalityKey'] = rel_df['Locality'].str.replace('_', '', regex=False).str.replace(' ', '', regex=False)
    rel_map = {(row['LocalityKey'], row['Pesticide']): float(row['rel_freq']) for _, row in rel_df.iterrows()}

    pair_src = OUTPUT_DIR / '19_pair_rules_topN_by_locality_report.xlsx'
    xls = pd.ExcelFile(pair_src)
    top20_src = OUTPUT_DIR / '20_pair_rules_top_heatmap_input.xlsx'
    with pd.ExcelWriter(top20_src, engine='openpyxl') as writer:
        for sheet in xls.sheet_names:
            df = pd.read_excel(pair_src, sheet_name=sheet).head(TOP_N_PAIR_HEATMAP).copy()
            locality = sheet
            df['Locality'] = locality
            df = df.rename(columns={'Pesticide A': 'Pesticide1', 'Pesticide B': 'Pesticide2', 'SupportRel_AB': 'SupportRel'})
            df.to_excel(writer, sheet_name=f'{sanitize_sheet(sheet)[:30]}2', index=False)

    xls2 = pd.ExcelFile(top20_src)
    for sheet in xls2.sheet_names:
        if not sheet.endswith('2'):
            continue
        df = pd.read_excel(top20_src, sheet_name=sheet)
        if not {'SupportRel', 'Pesticide1', 'Pesticide2', 'Locality'}.issubset(df.columns):
            continue
        df['SupportRel'] = pd.to_numeric(df['SupportRel'], errors='coerce')
        df = df.dropna(subset=['SupportRel'])
        loc = df['Locality'].iloc[0]
        loc_key = str(loc).replace('_', '').replace(' ', '')
        pesticides = sorted(set(df['Pesticide1']).union(set(df['Pesticide2'])))
        mat = pd.DataFrame(0.0, index=pesticides, columns=pesticides)
        for _, row in df.iterrows():
            a, b2, s = row['Pesticide1'], row['Pesticide2'], row['SupportRel']
            mat.loc[a, b2] = s
            mat.loc[b2, a] = s
        for p in pesticides:
            diag_val = rel_map.get((loc_key, p))
            if diag_val is not None:
                mat.loc[p, p] = diag_val
        fig, ax = plt.subplots(figsize=(max(8, len(pesticides) * 0.6), max(6, len(pesticides) * 0.55)))
        heatmap = sns.heatmap(mat, ax=ax, cmap='coolwarm', annot=True, fmt='.2f', vmin=0, vmax=1, square=True, linewidths=0.5, linecolor='white', cbar=False)
        divider = make_axes_locatable(ax)
        cax = divider.append_axes('right', size='3%', pad=0.12)
        cbar = fig.colorbar(heatmap.collections[0], cax=cax)
        cbar.outline.set_visible(False)
        cbar.set_label('Relative support')
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0)
        plt.tight_layout()
        safe_loc = sanitize_filename(sheet.replace('2', ''))
        plt.savefig(OUTPUT_DIR / f'21_pair_support_heatmap_{safe_loc}.pdf', bbox_inches='tight')
        plt.close()

    outdir = OUTPUT_DIR / '22_latex_pair_rule_tables'
    ensure_dir(outdir)
    col_headers = [r'\#', 'Pesticide A', 'Pesticide B', r'SuppRel$_{A}$', r'SuppRel$_{B}$', r'SuppRel$_{AB}$', r'Conf$_{A\to B}$', r'Conf$_{B\to A}$', 'Lift', r'Conv$_{A\to B}$', r'Conv$_{B\to A}$']
    def make_longtable(df, locality, label_key):
        ncols = len(col_headers)
        colspec = 'r l l ' + 'r ' * (ncols - 3)
        lines = []
        lines.append(r'\begin{longtable}{' + colspec.strip() + '}')
        lines.append(r'\caption{Top 50 pesticide-pair association rules for ' + esc_tex(locality) + r', ranked by relative support.}\label{tab:pairs_' + label_key + r'}\\')
        lines.append(r'\hline')
        lines.append(' & '.join(col_headers) + r'\\')
        lines.append(r'\hline')
        lines.append(r'\endfirsthead')
        lines.append(r'\multicolumn{' + str(ncols) + r'}{c}{-- continued --}\\')
        lines.append(r'\hline')
        lines.append(' & '.join(col_headers) + r'\\')
        lines.append(r'\hline')
        lines.append(r'\endhead')
        lines.append(r'\hline')
        lines.append(r'\endfoot')
        lines.append(r'\hline')
        lines.append(r'\endlastfoot')
        for _, row in df.iterrows():
            cells = [
                f"{int(row['RowIndex'])}", esc_tex(row['Pesticide A']), esc_tex(row['Pesticide B']),
                fmt_num(row['SupportRel_A']), fmt_num(row['SupportRel_B']), fmt_num(row['SupportRel_AB']),
                fmt_num(row['Confidence_A_to_B']), fmt_num(row['Confidence_B_to_A']), fmt_num(row['Lift']),
                fmt_num(row['Conviction_A_to_B']), fmt_num(row['Conviction_B_to_A'])
            ]
            lines.append(' & '.join(cells) + r'\\')
        lines.append(r'\end{longtable}')
        return '\n'.join(lines)

    all_locality_tex = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(pair_src, sheet_name=sheet)
        label_key = re.sub(r'[^a-zA-Z0-9]+', '', sheet.strip())
        texbody = make_longtable(df, sheet, label_key)
        standalone = '\n'.join([
            r'\documentclass{article}', r'\usepackage[margin=1cm]{geometry}', r'\usepackage{longtable}',
            r'\usepackage{pdflscape}', r'\begin{document}', r'\begin{landscape}', texbody, r'\end{landscape}', r'\end{document}'
        ])
        with open(outdir / f'top50pairs_{label_key}.tex', 'w', encoding='utf-8') as f:
            f.write(standalone)
        all_locality_tex[sheet] = texbody
    mainlines = [r'\documentclass{article}', r'\usepackage[margin=1cm]{geometry}', r'\usepackage{longtable}', r'\usepackage{pdflscape}', r'\begin{document}', r'\begin{landscape}']
    for _, texbody in all_locality_tex.items():
        mainlines.append(texbody)
        mainlines.append(r'\clearpage')
    mainlines += [r'\end{landscape}', r'\end{document}']
    with open(outdir / 'all_localities_top50.tex', 'w', encoding='utf-8') as f:
        f.write('\n'.join(mainlines))





def generate_association_rules_latex_tables():
    """Generate LaTeX longtable files for top association rules by locality."""
    src = OUTPUT_DIR / '28_association_rules_triplets_topN_by_locality.xlsx'
    outdir = OUTPUT_DIR / '29_latex_triplet_rule_tables'
    ensure_dir(outdir)
    xls = pd.ExcelFile(src)

    col_headers = [
        r'\#', 'RuleSize', 'Antecedent', 'Consequent', 'SuppRel', 'Confidence', 'Lift'
    ]

    def make_longtable(df, locality, label_key):
        ncols = len(col_headers)
        colspec = 'r r l l r r r'
        lines = []
        lines.append(r'\begin{longtable}{' + colspec + '}')
        lines.append(r'\caption{Top triplet association rules for ' + esc_tex(locality) + r', ranked by confidence.}\label{tab:assoc_' + label_key + r'}\\')
        lines.append(r'\hline')
        lines.append(' & '.join(col_headers) + r'\\')
        lines.append(r'\hline')
        lines.append(r'\endfirsthead')
        lines.append(r'\multicolumn{' + str(ncols) + r'}{c}{-- continued --}\\')
        lines.append(r'\hline')
        lines.append(' & '.join(col_headers) + r'\\')
        lines.append(r'\hline')
        lines.append(r'\endhead')
        lines.append(r'\hline')
        lines.append(r'\endfoot')
        lines.append(r'\hline')
        lines.append(r'\endlastfoot')
        for i, (_, row) in enumerate(df.iterrows(), start=1):
            cells = [
                str(i),
                str(int(row['RuleSize'])) if pd.notna(row['RuleSize']) else '',
                esc_tex(row['Antecedent']),
                esc_tex(row['Consequent']),
                fmt_num(row['SupportRel']),
                fmt_num(row['Confidence']),
                fmt_num(row['Lift'])
            ]
            lines.append(' & '.join(cells) + r'\\')
        lines.append(r'\end{longtable}')
        return '\n'.join(lines)

    all_tex = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(src, sheet_name=sheet)
        label_key = re.sub(r'[^a-zA-Z0-9]+', '', sheet.strip())
        texbody = make_longtable(df, sheet, label_key)
        standalone = '\n'.join([
            r'\documentclass{article}',
            r'\usepackage[margin=1cm]{geometry}',
            r'\usepackage{longtable}',
            r'\usepackage{pdflscape}',
            r'\begin{document}',
            r'\begin{landscape}',
            texbody,
            r'\end{landscape}',
            r'\end{document}'
        ])
        with open(outdir / f'top_triplet_rules_{label_key}.tex', 'w', encoding='utf-8') as f:
            f.write(standalone)
        all_tex[sheet] = texbody

    mainlines = [
        r'\documentclass{article}',
        r'\usepackage[margin=1cm]{geometry}',
        r'\usepackage{longtable}',
        r'\usepackage{pdflscape}',
        r'\begin{document}',
        r'\begin{landscape}'
    ]
    for _, texbody in all_tex.items():
        mainlines.append(texbody)
        mainlines.append(r'\clearpage')
    mainlines += [r'\end{landscape}', r'\end{document}']
    with open(outdir / 'all_localities_top_triplet_rules.tex', 'w', encoding='utf-8') as f:
        f.write('\n'.join(mainlines))


def merged_hamming_outputs(binary_sheets):
    """Compute merged-locality Hamming distances between pesticides and export plots.

    Parameters
    ----------
    binary_sheets : dict[str, pandas.DataFrame]
        Mapping from sheet name to binary pesticide-by-month matrix.
    """
    merged = pd.concat(list(binary_sheets.values()), axis=1)
    X = merged.values.astype(int)
    labels = merged.index.astype(str).tolist()
    D = squareform(pdist(X, metric='hamming'))
    D_df = pd.DataFrame(D, index=labels, columns=labels)
    csv_path = OUTPUT_DIR / '23_hamming_distance_matrix_all_localities_merged.csv'
    D_df.to_csv(csv_path, sep=';')


    fig, ax = plt.subplots(figsize=(18, 24))
    heatmap = sns.heatmap(D_df, ax=ax, cmap='viridis', square=True, cbar=False)
    divider = make_axes_locatable(ax)
    cax = divider.append_axes('right', size='3%', pad=0.3)
    cbar = fig.colorbar(heatmap.collections[0], cax=cax)
    cbar.ax.tick_params(labelsize=13)
    cbar.set_label('Hamming distance', fontsize=13)
    ax.set_xticklabels(ax.get_xticklabels(), rotation=90, fontsize=11)
    ax.set_yticklabels(ax.get_yticklabels(), rotation=0, fontsize=11)
    n = D_df.shape[0]
    for i in range(10, n, 10):
        ax.axhline(i, color='white', linewidth=1.2)
        ax.axvline(i, color='white', linewidth=1.2)
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / '24_hamming_distance_heatmap_grid10.pdf', format='pdf', dpi=300, bbox_inches='tight')
    plt.close()

    Z = linkage(pdist(X, metric='hamming'), method='average')
    plt.figure(figsize=(10, 14))
    dendrogram(Z, labels=labels, orientation='right', leaf_font_size=6)
    plt.xlabel('Hamming distance (proportion of differing columns)')
    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / '25_hamming_distance_dendrogram.pdf', format='pdf')
    plt.close()
    return D_df


def main():
    """Run the full pesticide analysis pipeline from workbook to final outputs."""
    total_start = time.perf_counter()

    step_start = time.perf_counter()
    sheets = load_workbook_sheets(INPUT_XLSX)
    print(f"Step load_workbook_sheets: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    binary_sheets, _ = build_binary_outputs(sheets)
    print(f"Step build_binary_outputs: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    compute_tsne(binary_sheets)
    print(f"Step compute_tsne: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    relfreq_df, _ = relative_frequency_outputs(sheets)
    print(f"Step relative_frequency_outputs: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    pair_and_triplet_itemsets(sheets)
    print(f"Step pair_and_triplet_itemsets: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    compute_pair_itemsets_confidence(sheets)
    print(f"Step compute_pair_itemsets_confidence: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    compute_association_rules_triplets(sheets, min_support=MIN_SUPPORT, min_confidence=MIN_CONFIDENCE)
    print(f"Step compute_association_rules_triplets: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    generate_association_rules_latex_tables()
    print(f"Step generate_association_rules_latex_tables: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    pair_heatmaps_and_latex(relfreq_df)
    print(f"Step pair_heatmaps_and_latex: {time.perf_counter() - step_start:.3f} s")

    step_start = time.perf_counter()
    merged_hamming_outputs(binary_sheets)
    print(f"Step merged_hamming_outputs: {time.perf_counter() - step_start:.3f} s")

    total_elapsed = time.perf_counter() - total_start
    print(f"Total pipeline time: {total_elapsed:.3f} s")
    print('Pipeline completed. Outputs written to output/.')


if __name__ == '__main__':
    main()
